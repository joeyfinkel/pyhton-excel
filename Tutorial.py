import json
import itertools
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

grades_book = load_workbook('Grades.xlsx')
new_grades_book = load_workbook('New Grades.xlsx')
ws = grades_book.active
ws2 = new_grades_book.active


def get_headers(ws):
    value = []
    row = [i[0] for i in ws.iter_cols(values_only=True)]

    # Deconstruct list
    for i in row:
        value += [i]

    return value

def populate_col(sheet, total_columns):
    print('NEEDS TO BE COMPLETED')

def get_column(column, sheet, row) -> list:
    col = []

    for i in sheet.iter_rows(min_row = row, values_only = True):
        col.append(i[column])

    return col

def insert_every_n(l1, l2, k):
    i1, i2 = iter(l1), iter(l2)
    while True:
        try:
            yield from itertools.islice(i1, k)
            yield next(i2)
        except StopIteration:
            return

# create the list of headings for amount of rows in the column
heading = get_column(1, ws, 1)[0]
head = [heading] * len(get_column(1, ws, 2))

# create list of names
name = get_column(1, ws, 2)
new_list = list(insert_every_n(head, name, k=1)) # combine the list of names and headings

results = []
values = {}
for key, value in zip(new_list[0::2], new_list[1::2]):
    values = dict(zip([key], [value]))
    results.append(values)

d = []
data = {}
for result in results:
    print(get_column(0, ws, 2))
    # data = dict(zip([i for i in get_column(0, ws, 2)], [result]))
    # data = {i:result for i in get_column(0, ws, 2)}

    # print(result)

# data = {i:result for i in get_column(0, ws, 2)}
    # print(data)
# headings = ['ID'] + list(data[1].keys())
# ws2.append(headings)
# print(names)
# print(headings)

# for person in data:
#     name = list(data[person].values())
#     ws2.append([person] + name)


# print(names)
# populate_col(ws, 2)
# for c in get_column(0):
#     r = 1
#     ws.cell(row = r + 1, column = 1).value = c
# new_grades_book.save('New Grades.xlsx')
# umn):
#     ws2.cell(row = 2, column = 1, value = 'HELLO')\